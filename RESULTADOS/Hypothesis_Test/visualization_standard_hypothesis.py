# -*- coding: utf8 -*-

import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill
import math
import pywt
import csv

# CONFIGURAÇÕES - ARQUIVO EXCEL
# A -> W+RFC            B -> W+1NN          C -> W+GNB          D -> W + SVM
# E -> W+PCA+RFC        F -> W+PCA+1NN      G -> W+PCA+GNB      H -> W+PCA+SVM
# I -> W+LDA+RFC        J -> W+LDA+1NN      K -> W+LDA+GNB      L -> W+LDA+SVM
# M -> W+PCA+LDA+RFC    N -> W+PCA+LDA+1NN  O -> W+PCA+LDA+GNB  P -> W+PCA+LDA+SVM

if __name__ == "__main__":
    databases = ["AR", "YaleB", "ORL", "GTech", "Essex"]
    levels = ["Level1", "Level2", "Level3", "Level4", "Level5"]
    workbooks = ["C:\\Users\\Fabricio\\PycharmProjects\\Waveletfaces\\RESULTADOS\\resultados_waveletfaces_AR.xlsx",
                 "C:\\Users\\Fabricio\\PycharmProjects\\Waveletfaces\\RESULTADOS\\resultados_waveletfaces_YaleB.xlsx",
                 "C:\\Users\\Fabricio\\PycharmProjects\\Waveletfaces\\RESULTADOS\\resultados_waveletfaces_ORL.xlsx",
                 "C:\\Users\\Fabricio\\PycharmProjects\\Waveletfaces\\RESULTADOS\\resultados_waveletfaces_GTech.xlsx",
                 "C:\\Users\\Fabricio\\PycharmProjects\\Waveletfaces\\RESULTADOS\\resultados_waveletfaces_Essex.xlsx"]
    
    blackFill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    grayFill = PatternFill(start_color='A8A8A8', end_color='A8A8A8', fill_type="solid")

    waves = pywt.wavelist(kind="discrete")  # wavelets

    for i in range(0, len(databases)):
        count, classifier, config = 0, 0, -1
        mean, std = 0, 0
        
        if i == 0:    best_mean, best_std = 97.9292700000001, 0.380296743652012  # AR - Level 3 - W+LDA+1NN (rbio3.1)                  
        elif i == 1:  best_mean, best_std = 93.88572, 0.710512733911577          # YaleB - Level 4 - W+LDA+1NN (rbio4.4)
        elif i == 2:  best_mean, best_std = 97.525, 1.17502417770096             # ORL - Level 2 - W+PCA+LDA+1NN (rbio2.6)                HIGHEST ACCURACY OF WAVELETS
        elif i == 3:  best_mean, best_std = 84.82, 1.70230087905408              # GTech - Level 5 - W+LDA+1NN (rbio3.1)
        elif i == 4:  best_mean, best_std = 95.19862, 0.849234323140558          # Essex - Level 5 - W+PCA+LDA+1NN (rbio3.1)
        
        wb = load_workbook(filename=workbooks[i])
        wb2 = Workbook()
        
        for level in levels:
            wavelet = ""
            sign = False
            table = wb2.create_sheet(title=databases[i] + "_" + level)
            sheet = wb[databases[i] + "_" + level]

            for row in sheet.rows:
                for cell in row:
                    if cell.value == None or cell.value == "-": continue
                    count += 1
                    if count <= 20: continue
                    if isinstance(cell.value, basestring):
                        if cell.value == wavelet: config += 1
                        else: config = 0
                        wavelet = cell.value
                        classifier = 0
                    else:
                        if sign == True:
                            sign = False
                            std = cell.value
                            interval_max = (best_mean - mean) + 1.96 * math.sqrt(best_std ** 2 + std ** 2)
                            interval_min = (best_mean - mean) - 1.96 * math.sqrt(best_std ** 2 + std ** 2)
                            if interval_min <= 0 <= interval_max:       # Does the interval contain 0?
                                if config == 0:
                                    if classifier == 0: col = "A"
                                    if classifier == 1: col = "B"
                                    if classifier == 2: col = "C"
                                    if classifier == 3: col = "D"
                                elif config == 1:
                                    if classifier == 0: col = "E"
                                    if classifier == 1: col = "F"
                                    if classifier == 2: col = "G"
                                    if classifier == 3: col = "H"
                                elif config == 2:
                                    if classifier == 0: col = "I"
                                    if classifier == 1: col = "J"
                                    if classifier == 2: col = "K"
                                    if classifier == 3: col = "L"
                                elif config == 3:
                                    if classifier == 0: col = "M"
                                    if classifier == 1: col = "N"
                                    if classifier == 2: col = "O"
                                    if classifier == 3: col = "P"
                                table[col + str(waves.index(wavelet) + 1)].fill = grayFill
                            classifier += 1
                        else:
                            mean = cell.value                    
                            sign = True
        wb2.save(databases[i] + ".xlsx")                
